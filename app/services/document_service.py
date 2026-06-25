from io import BytesIO
import logging
import pdfplumber
from docx import Document as DocxDocument
import pytesseract
from pdf2image import convert_from_bytes
import openpyxl
from PIL import Image

logger = logging.getLogger(__name__)


def extract_text(file_bytes: bytes, file_type: str) -> str:
    if file_type in {"txt", "md"}:
        return file_bytes.decode("utf-8", errors="ignore")
    if file_type == "pdf":
        return _extract_pdf_text(file_bytes)
    if file_type == "docx":
        return _extract_docx_text(file_bytes)
    if file_type in {"xlsx", "xls"}:
        return _extract_xlsx_text(file_bytes)
    if file_type in {"png", "jpg", "jpeg"}:
        return _extract_image_text(file_bytes)
    raise ValueError(f"Unsupported file type: {file_type}")


def _extract_pdf_text(file_bytes: bytes) -> str:
    parts: list[str] = []
    try:
        with pdfplumber.open(BytesIO(file_bytes)) as pdf:
            for idx, page in enumerate(pdf.pages):
                page_num = idx + 1
                txt = (page.extract_text() or "").strip()
                # Check if this page contains meaningful text or is scanned
                if len(txt) > 10:
                    logger.info("Page %d has text content. Extracting as-is.", page_num)
                    parts.append(txt)
                else:
                    logger.info("Page %d has insufficient text (%d chars). Falling back to OCR...", page_num, len(txt))
                    try:
                        images = convert_from_bytes(file_bytes, first_page=page_num, last_page=page_num)
                        if images:
                            page_text = pytesseract.image_to_string(images[0]).strip()
                            parts.append(page_text)
                        else:
                            parts.append(txt)
                    except Exception as ocr_err:
                        logger.error("OCR failed for page %d: %s", page_num, ocr_err)
                        parts.append(txt)
    except Exception as e:
        logger.warning("PDF parsing failed: %s. Attempting full OCR fallback.", e)
        # Full fallback if the entire pdf structure is corrupt or cannot be opened by pdfplumber
        try:
            images = convert_from_bytes(file_bytes)
            for i, image in enumerate(images):
                logger.info("OCR-ing page %d...", i + 1)
                page_text = pytesseract.image_to_string(image)
                parts.append(page_text)
        except Exception as e_full:
            logger.error("Full OCR fallback failed: %s", e_full)
            
    return "\n".join(parts).strip()


def _extract_docx_text(file_bytes: bytes) -> str:
    document = DocxDocument(BytesIO(file_bytes))
    return "\n".join(paragraph.text for paragraph in document.paragraphs).strip()


def _extract_xlsx_text(file_bytes: bytes) -> str:
    parts = []
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            parts.append(f"--- Sheet: {sheet} ---")
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    parts.append(row_text)
    except Exception as e:
        logger.error("Excel extraction failed: %s", e)
    return "\n".join(parts).strip()


def _extract_image_text(file_bytes: bytes) -> str:
    try:
        image = Image.open(BytesIO(file_bytes))
        return pytesseract.image_to_string(image).strip()
    except Exception as e:
        logger.error("Image OCR failed: %s", e)
        return ""
